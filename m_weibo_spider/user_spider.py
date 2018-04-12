# -*- coding: utf-8 -*-
"""
Created on Sat Mar 17 18:42:11 2018
爬去指定用户的已发布微博
    args：用户主页URL
@author: fivelike

微博用户主页url有两种形式：
    1,https://m.weibo.cn/p/100505 + oid
    2,https://m.weibo.cn/u/oid
    (oid为用户唯一标识符)
获取Json数据API有两种:
    1,https://m.weibo.cn/api/container/getIndex?containerid=100505+oid 获取用户信息
    2,https://m.weibo.cn/api/container/getIndex?containerid=107603+oid 获取已发布微博列表
    分页：https://m.weibo.cn/api/container/getIndex?containerid=107603+oid & page=2(3,4,5...)
    
    无数据 "ok":"0"
"""

import requests
import json
import openpyxl
import time

class UserByUrl:
    session = requests.Session()
    oid = ""
    
    def __init__(self,url):
        if "https://m.weibo.cn/p" in url:
            oid = url.replace("https://m.weibo.cn/p/100505","") 
        elif "https://m.weibo.cn/u" in url:
            oid = url.replace("https://m.weibo.cn/u/","")
        else:
            print("url错误！")
        if oid:
            self.userinfo_url = "https://m.weibo.cn/api/container/getIndex?containerid=100505"+oid
            self.content_url = "https://m.weibo.cn/api/container/getIndex?containerid=107603"+oid     
#        print(self.userinfo_url)
#        print(self.content_url)

    def get_user_data(self):
        try:
            info_res = self.session.get(self.userinfo_url)
        except:
            print("获取用户基本信息失败！")
        info_data = json.loads(info_res.text)
        username = info_data["data"]["userInfo"]["screen_name"]
#        print(username)
        return username
    
    def write_content_data(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "微博信息"
        ws1.append(["文章id","微博内容","发布时间","转发数","点赞数","评论数"])
        
        wb.create_sheet("评论信息",1).append(["文章id","评论内容","点赞数","评论用户名","用户id","发布时间"])
        wb.create_sheet("转发信息",2).append(["文章id","转发附加文字","转发者用户名","用户id","转发时间"])
        
        count = 0
        
        for i in range(2, 1000):
            try:
                content_res = self.session.get(self.content_url+"&page="+str(i))
            except:
                print("获取page"+str(i)+"失败")
                continue
            json_data = json.loads(content_res.text)
            if json_data["ok"] == 0:
                break
            cards = json_data["data"]["cards"]
            index = 0
            for card in cards:
#                print("=============")
#                print(card["mblog"]["text"])
#                 print("发布时间:"+card["mblog"]["created_at"])
#                print("转发数:"+str(card["mblog"]["reposts_count"]))
#                print("点赞数:"+str(card["mblog"]["attitudes_count"]))
#                print("评论数:"+str(card["mblog"]["comments_count"]))
#                print("详情页Url:"+ "https://m.weibo.cn/status/"+str(card["mblog"]["id"]))
                
                if card["mblog"]["created_at"].find("-") > -1:
                     tempTime = "2018-" + card["mblog"]["created_at"] + " 00:00:01"
                     createTime = time.mktime(time.strptime(tempTime, "%Y-%m-%d %H:%M:%S"))
                     endTime = time.mktime(time.strptime("2018-01-31 23:59:59", "%Y-%m-%d %H:%M:%S"))
                     if createTime < endTime:
                         
                         print("完了？")
                         break
                text = card["mblog"]["text"]
                created_at = card["mblog"]["created_at"]
                reposts_count = str(card["mblog"]["reposts_count"])
                attitudes_count = str(card["mblog"]["attitudes_count"])
                comments_count = str(card["mblog"]["comments_count"])
                weibo_id = str(card["mblog"]["id"])
                index = index + 1
                try:
                    ws1.append([weibo_id,
                               text,
                               created_at,
                               reposts_count,
                               attitudes_count,
                               comments_count
                               ])
                    count+=1
                    if count%200==0:
                        print("已写入{0}条微博数据".format(count))
                except:
                    print("插入数据失败！")
                    continue
            if index < len(cards):
                print("操")
                break
        #保存文件
        username = self.get_user_data()
        wb.save(username + ".xlsx")
        print("共爬取用户：{0}——{1}条微博".format(username,count))
        
                
